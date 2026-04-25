"""
nghiepvu/views.py

Quyền:
  - Điểm danh (ăn/ngủ)    : admin, hoc_vu
  - Xuất file (API data)   : admin, hoc_vu
  - Lịch trực xem          : tất cả (login required)
  - Phân công trực         : admin, quan_ly
  - Báo cáo/thống kê       : tất cả (login required)
"""
import json
from datetime import date, datetime, timedelta

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.db.models import Count, Q
from django.views.decorators.http import require_POST

from accounts.decorators import role_required
from quanli.models import HocSinh, GiaoVien, Phong
from core.models import LoaiTruc, TrangThaiDiemDanh, CauHinhHeThong, CauHinhGia
from .models import DiemDanhHS, DiemDanhPhong, PhanCongTrucGV
from .forms import PhanCongTrucGVForm


# ─────────────────────────────────────────────────────────
# ĐIỂM DANH ĂN
# ─────────────────────────────────────────────────────────
@role_required('admin', 'hoc_vu')
def diemdanh_an(request):
    return render(request, 'nghiepvu/diemdanh_an.html', {'active': 'diemdanh_an'})


@role_required('admin', 'hoc_vu')
def diemdanh_ngu(request):
    return render(request, 'nghiepvu/diemdanh_ngu.html', {'active': 'diemdanh_ngu'})


# ─────────────────────────────────────────────────────────
# API: Danh sách phòng & học sinh (cho JS fetch)
# ─────────────────────────────────────────────────────────
@role_required('admin', 'hoc_vu')
def api_phong_list(request, loai):
    """
    GET /api/phong/<loai>/   loai: 'an' | 'ngu'
    Trả về danh sách phòng theo loại.
    """
    loai_int = LoaiTruc.AN if loai == 'an' else LoaiTruc.NGU
    ds = list(
        Phong.objects.filter(loai_phong=loai_int)
        .values('ma_phong', 'suc_chua', 'gioi_tinh')
        .order_by('ma_phong')
    )
    return JsonResponse({'phong': ds})


@role_required('admin', 'hoc_vu')
def api_hocsinh_list(request, loai):
    """
    GET /api/hocsinh/<loai>/    loai: 'an' | 'ngu'
    Trả về toàn bộ HS đang học kèm thông tin phòng tương ứng.
    """
    qs = HocSinh.objects.filter(dang_hoc=True).select_related('ma_phong_an', 'ma_phong_ngu')

    if loai == 'an':
        qs = qs.filter(ma_phong_an__isnull=False).order_by('ma_phong_an', 'lop', 'ho_ten')
    else:
        qs = qs.filter(ma_phong_ngu__isnull=False).order_by('ma_phong_ngu', 'lop', 'ho_ten')

    def _khoi(lop):
        """Lấy khối từ tên lớp. VD: '10A1' → 10, '11B2' → 11"""
        try: return int(str(lop)[:2])
        except: return 0

    data = []
    for hs in qs:
        phong_an  = hs.ma_phong_an.ma_phong  if hs.ma_phong_an  else None
        phong_ngu = hs.ma_phong_ngu.ma_phong if hs.ma_phong_ngu else None
        data.append({
            'id':        hs.pk,
            'ma_so_bt':  hs.pk,  # Dùng id làm mã BT
            'ho_ten':    hs.ho_ten,
            'lop':       hs.lop,
            'khoi':      _khoi(hs.lop),
            'gioi_tinh': hs.gioi_tinh,
            'phong_an':  phong_an,
            'phong_ngu': phong_ngu,
        })
    return JsonResponse({'hocsinh': data})


@role_required('admin', 'hoc_vu')
def api_diemdanh_get(request):
    """
    GET /api/diemdanh/?ngay=YYYY-MM-DD&loai=an|ngu
    Trả về dữ liệu điểm danh đã lưu cho ngày đó.
    """
    ngay_str = request.GET.get('ngay')
    loai     = request.GET.get('loai', 'an')
    if not ngay_str:
        return JsonResponse({'error': 'Thiếu tham số ngay'}, status=400)

    try:
        ngay = datetime.strptime(ngay_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Ngày không hợp lệ'}, status=400)

    qs = DiemDanhHS.objects.filter(ngay=ngay).values(
        'ma_hs_id', 'diem_danh_an', 'diem_danh_ngu', 'ghi_chu'
    )

    field = 'diem_danh_an' if loai == 'an' else 'diem_danh_ngu'
    result = {
        row['ma_hs_id']: {
            'status': row[field],
            'ghi_chu': row['ghi_chu'] or '',
        }
        for row in qs
    }
    return JsonResponse({'records': result})


@require_POST
@role_required('admin', 'hoc_vu')
def api_diemdanh_save(request):
    """
    POST /api/diemdanh/save/
    Body JSON: { loai: 'an'|'ngu', records: [{ma_hs, ngay, status, ghi_chu}, ...] }
    Dùng upsert (update_or_create) để lưu.
    """
    try:
        body  = json.loads(request.body)
        loai  = body.get('loai', 'an')
        recs  = body.get('records', [])
    except (json.JSONDecodeError, KeyError):
        return JsonResponse({'error': 'Dữ liệu không hợp lệ'}, status=400)

    field = 'diem_danh_an' if loai == 'an' else 'diem_danh_ngu'
    saved = 0

    with transaction.atomic():
        for rec in recs:
            try:
                ma_hs = int(rec['ma_hs'])
                ngay  = datetime.strptime(rec['ngay'], '%Y-%m-%d').date()
                status = int(rec.get('status', TrangThaiDiemDanh.CO_MAT))
                ghi_chu = rec.get('ghi_chu', '') or ''

                hs     = HocSinh.objects.get(pk=ma_hs, dang_hoc=True)
                obj, _ = DiemDanhHS.objects.get_or_create(
                    ma_hs=hs, ngay=ngay,
                    defaults={
                        'diem_danh_an': TrangThaiDiemDanh.CO_MAT,
                        'diem_danh_ngu': TrangThaiDiemDanh.CO_MAT,
                    }
                )
                setattr(obj, field, status)
                obj.ghi_chu = ghi_chu
                obj.save()
                saved += 1
            except (HocSinh.DoesNotExist, ValueError, KeyError):
                continue

    return JsonResponse({'saved': saved, 'ok': True})


# ─────────────────────────────────────────────────────────
# LỊCH TRỰC GV
# ─────────────────────────────────────────────────────────
@login_required(login_url='/login/')
def lichtruc(request):
    """Xem lịch trực – tất cả role."""
    ngay_str = request.GET.get('ngay', date.today().isoformat())
    try:
        ngay = datetime.strptime(ngay_str, '%Y-%m-%d').date()
    except ValueError:
        ngay = date.today()

    qs = PhanCongTrucGV.objects.filter(ngay=ngay).select_related(
        'ma_gv', 'ma_gv_truc_thay', 'ma_phong'
    ).order_by('loai_truc', 'ma_phong__ma_phong')

    return render(request, 'nghiepvu/lichtruc.html', {
        'ds_phan_cong': qs,
        'ngay': ngay,
        'active': 'lichtruc',
    })


@role_required('admin', 'quan_ly')
def phan_cong_truc(request):
    tuan_str = request.GET.get('tuan')
    try:
        mon = datetime.strptime(tuan_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        today = date.today()
        mon = today - timedelta(days=today.weekday())
    fri = mon + timedelta(days=4)
    qs = PhanCongTrucGV.objects.filter(ngay__gte=mon, ngay__lte=fri).select_related('ma_gv','ma_phong','ma_gv_truc_thay')
    pc_data = []
    for pc in qs:
        pc_data.append({'id':pc.pk,'ma_gv_id':pc.ma_gv_id,'ma_gv_truc_thay_id':pc.ma_gv_truc_thay_id,'ma_phong_id':pc.ma_phong.ma_phong,'ngay':pc.ngay.isoformat(),'loai_truc':pc.loai_truc,'xac_nhan_truc':pc.xac_nhan_truc})
    gv_list = list(GiaoVien.objects.filter(dang_lam=True).values('id','ho_ten','gioi_tinh','lich_ranh'))
    phong_list = list(Phong.objects.values('ma_phong','loai_phong','suc_chua','gioi_tinh'))
    return render(request, 'nghiepvu/lichtruc_admin.html', {'active': 'admin_lichtruc', 'pc_json': json.dumps(pc_data), 'gv_json': json.dumps(gv_list), 'phong_json': json.dumps(phong_list), 'tuan_bat_dau': mon.isoformat()})

@role_required('admin', 'quan_ly')
def phan_cong_xoa(request, pk):
    pc = get_object_or_404(PhanCongTrucGV, pk=pk)
    if request.method == 'POST':
        pc.delete()
        messages.success(request, 'Đã xóa phân công.')
        return redirect('admin_lichtruc')
    return render(request, 'core/confirm_delete.html', {
        'ten': f"Phân công {pc.ma_gv} – {pc.ngay}"
    })


# ─────────────────────────────────────────────────────────
# BÁO CÁO / THỐNG KÊ
# ─────────────────────────────────────────────────────────
@login_required(login_url='/login/')
def baocao(request):
    """Xem báo cáo – tất cả role."""
    return render(request, 'nghiepvu/baocao.html', {'active': 'baocao'})


# ─────────────────────────────────────────────────────────
# API BÁO CÁO: Danh sách điểm danh theo phòng (xuất file)
# ─────────────────────────────────────────────────────────
@login_required(login_url='/login/')
def api_baocao_diemdanh(request):
    """
    GET /api/baocao/diemdanh/?loai=an&thang=3&nam=2026&lop=10A1
    Trả về danh sách học sinh đã điểm danh trong tháng.
    """
    loai     = request.GET.get('loai', 'an')
    thang    = int(request.GET.get('thang', date.today().month))
    nam      = int(request.GET.get('nam', date.today().year))
    lop_filter = request.GET.get('lop', '').strip()

    qs = DiemDanhHS.objects.filter(
        ngay__year=nam, ngay__month=thang
    ).select_related('ma_hs', 'ma_hs__ma_phong_an', 'ma_hs__ma_phong_ngu')

    if lop_filter:
        qs = qs.filter(ma_hs__lop__iexact=lop_filter)

    field = 'diem_danh_an' if loai == 'an' else 'diem_danh_ngu'

    data = []
    for rec in qs.order_by('ma_hs__ma_phong_an', 'ma_hs__lop', 'ma_hs__ho_ten', 'ngay'):
        phong = rec.ma_hs.ma_phong_an.ma_phong if loai == 'an' else rec.ma_hs.ma_phong_ngu.ma_phong
        data.append({
            'id':      rec.ma_hs.pk,
            'ho_ten':  rec.ma_hs.ho_ten,
            'lop':     rec.ma_hs.lop,
            'phong':   phong,
            'ngay':    rec.ngay.isoformat(),
            'status':  getattr(rec, field),
            'ghi_chu': rec.ghi_chu or '',
        })

    ds_lop = list(
        HocSinh.objects.filter(dang_hoc=True)
        .values_list('lop', flat=True).distinct().order_by('lop')
    )

    return JsonResponse({'records': data, 'ds_lop': ds_lop})


@login_required(login_url='/login/')
def api_baocao_full(request):
    """
    GET /api/baocao/full/
    Trả về:
      - students: tất cả HS đang học (id, ho_ten, lop, khoi, phong_an, phong_ngu)
      - records:  tất cả DiemDanhHS (ma_hs, ngay, diem_danh_an, diem_danh_ngu)
      - phong_an_list, phong_ngu_list: ds phòng cho bộ lọc
      - ds_lop: ds lớp cho bộ lọc
    """
    def _khoi(lop):
        try:
            return str(lop)[:2]
        except:
            return ''

    # -- Students --
    students = []
    for hs in HocSinh.objects.filter(dang_hoc=True).select_related('ma_phong_an', 'ma_phong_ngu').order_by('lop', 'ho_ten'):
        students.append({
            'id':       hs.pk,
            'ho_ten':   hs.ho_ten,
            'lop':      hs.lop,
            'khoi':     _khoi(hs.lop),
            'phong_an': hs.ma_phong_an.ma_phong if hs.ma_phong_an else '',
            'phong_ngu': hs.ma_phong_ngu.ma_phong if hs.ma_phong_ngu else '',
        })

    # -- Attendance records --
    records = []
    for rec in DiemDanhHS.objects.all().order_by('ngay'):
        records.append({
            'ma_hs':         rec.ma_hs_id,
            'ngay':          rec.ngay.isoformat(),
            'diem_danh_an':  rec.diem_danh_an,
            'diem_danh_ngu': rec.diem_danh_ngu,
        })

    # -- Room lists for filters --
    phong_an_list = list(
        Phong.objects.filter(loai_phong=LoaiTruc.AN)
        .values_list('ma_phong', flat=True).order_by('ma_phong')
    )
    phong_ngu_list = list(
        Phong.objects.filter(loai_phong=LoaiTruc.NGU)
        .values_list('ma_phong', flat=True).order_by('ma_phong')
    )
    ds_lop = list(
        HocSinh.objects.filter(dang_hoc=True)
        .values_list('lop', flat=True).distinct().order_by('lop')
    )

    return JsonResponse({
        'students':      students,
        'records':       records,
        'phong_an_list': phong_an_list,
        'phong_ngu_list': phong_ngu_list,
        'ds_lop':        ds_lop,
    })


@login_required(login_url='/login/')
def api_baocao_luong_gv(request):
    """
    GET /api/baocao/luong-gv/?thang=3&nam=2026
    Trả về danh sách tiền công GV trong tháng.
    (Mỗi buổi trực = 1 ca; giá = CauHinhGia.don_gia tương ứng loại trực)
    """
    thang = int(request.GET.get('thang', date.today().month))
    nam   = int(request.GET.get('nam', date.today().year))

    qs = PhanCongTrucGV.objects.filter(
        ngay__year=nam, ngay__month=thang, xac_nhan_truc=True
    ).select_related('ma_gv', 'ma_phong')

    # Lấy đơn giá mới nhất theo loại trực
    gia_map = {}
    for loai in [LoaiTruc.AN, LoaiTruc.NGU]:
        cfg = CauHinhGia.objects.filter(
            loai_truc=loai, ngay_ap_dung__lte=date(nam, thang, 28)
        ).order_by('-ngay_ap_dung').first()
        gia_map[loai] = float(cfg.don_gia) if cfg else 0

    # Gom theo GV
    gv_data = {}
    for pc in qs:
        gid = pc.ma_gv.pk
        if gid not in gv_data:
            gv_data[gid] = {
                'ho_ten':   pc.ma_gv.ho_ten,
                'so_ca_an': 0,
                'so_ca_ngu': 0,
                'tong_tien': 0.0,
            }
        ca = LoaiTruc.AN if pc.loai_truc == LoaiTruc.AN else LoaiTruc.NGU
        if ca == LoaiTruc.AN:
            gv_data[gid]['so_ca_an'] += 1
        else:
            gv_data[gid]['so_ca_ngu'] += 1
        gv_data[gid]['tong_tien'] += gia_map.get(ca, 0)

    result = sorted(gv_data.values(), key=lambda x: x['ho_ten'])
    return JsonResponse({
        'records': result,
        'gia_an':  gia_map.get(LoaiTruc.AN, 0),
        'gia_ngu': gia_map.get(LoaiTruc.NGU, 0),
    })


@require_POST
@role_required('admin', 'quan_ly')
def api_lichtruc_save(request):
    try:
        data = json.loads(request.body)
        pid = data.get('id')
        if pid:
            pc = get_object_or_404(PhanCongTrucGV, pk=pid)
        else:
            pc = PhanCongTrucGV()
        
        pc.ma_gv_id = data.get('ma_gv_id')
        pc.ma_gv_truc_thay_id = data.get('ma_gv_truc_thay_id')
        pc.ma_phong = get_object_or_404(Phong, ma_phong=data.get('ma_phong_id'))
        pc.ngay = data.get('ngay')
        pc.loai_truc = data.get('loai_truc')
        pc.save()
        
        return JsonResponse({'ok': True, 'id': pc.pk})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@require_POST
@role_required('admin', 'quan_ly')
def api_lichtruc_delete(request):
    try:
        data = json.loads(request.body)
        pid = data.get('id')
        pc = get_object_or_404(PhanCongTrucGV, pk=pid)
        pc.delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)

# ==========================================
# LỊCH KHUNG (CỐ ĐỊNH) - ENDPOINTS
# ==========================================
from nghiepvu.models import LichTrucCoDinh, ThuTrongTuan
import random

@role_required('admin', 'quan_ly')
def phan_cong_khung(request):
    qs = LichTrucCoDinh.objects.all().select_related('ma_phong', 'ma_gv')
    data = []
    for item in qs:
        data.append({
            'id': item.id,
            'ma_phong': item.ma_phong_id,
            'ma_gv': item.ma_gv_id,
            'thu': item.thu
        })
    
    phong_list = list(Phong.objects.values('ma_phong', 'loai_phong', 'suc_chua', 'gioi_tinh', 'sl_diem_danh', 'sl_ho_tro'))
    gv_list = list(GiaoVien.objects.filter(dang_lam=True).values('id', 'ho_ten', 'gioi_tinh', 'nhiem_vu', 'lich_ranh'))
    for g in gv_list:
        if not isinstance(g['lich_ranh'], list) or len(g['lich_ranh']) < 5:
            g['lich_ranh'] = [False, False, False, False, False]
    
    return render(request, 'nghiepvu/lichtruc_khung.html', {
        'active': 'admin_lichtruc_khung',
        'khung_json': json.dumps(data),
        'phong_json': json.dumps(phong_list),
        'gv_json': json.dumps(gv_list)
    })

@require_POST
@role_required('admin', 'quan_ly')
def api_lichtruc_khung_auto(request):
    try:
        with transaction.atomic():
            LichTrucCoDinh.objects.all().delete()
            
            # Lấy tất cả GV đang làm, tách theo nhiệm vụ
            gv_diem_danh = list(GiaoVien.objects.filter(dang_lam=True, nhiem_vu=0))
            gv_ho_tro    = list(GiaoVien.objects.filter(dang_lam=True, nhiem_vu=1))
            
            phongs_an  = [p for p in Phong.objects.all() if p.loai_phong == LoaiTruc.AN]
            phongs_ngu = [p for p in Phong.objects.all() if p.loai_phong == LoaiTruc.NGU]
            
            for thu, label in ThuTrongTuan.choices:
                # Pool GV rảnh vào ngày này
                pool_dd = [g for g in gv_diem_danh if isinstance(g.lich_ranh, list) and len(g.lich_ranh) > thu and g.lich_ranh[thu]]
                pool_ht = [g for g in gv_ho_tro    if isinstance(g.lich_ranh, list) and len(g.lich_ranh) > thu and g.lich_ranh[thu]]
                random.shuffle(pool_dd)
                random.shuffle(pool_ht)
                
                # Theo dõi ai đã được xếp Ăn / Ngủ riêng biệt
                assigned_an  = set()   # GV id đã xếp phòng Ăn hôm này
                assigned_ngu = set()   # GV id đã xếp phòng Ngủ hôm này
                
                def _pick(pool, assigned_set, req_gender):
                    """Lấy 1 GV từ pool, chưa được xếp loại này, khớp giới tính."""
                    for i, g in enumerate(pool):
                        if g.id in assigned_set:
                            continue
                        if req_gender is not None and g.gioi_tinh != req_gender:
                            continue
                        assigned_set.add(g.id)
                        return g
                    return None
                
                # ── Phòng Ăn (không phân biệt giới tính) ──
                for p in phongs_an:
                    for _ in range(p.sl_diem_danh):
                        g = _pick(pool_dd, assigned_an, None)
                        if g:
                            LichTrucCoDinh.objects.create(ma_phong=p, ma_gv=g, thu=thu)
                    for _ in range(p.sl_ho_tro):
                        g = _pick(pool_ht, assigned_an, None)
                        if g:
                            LichTrucCoDinh.objects.create(ma_phong=p, ma_gv=g, thu=thu)
                
                # ── Phòng Ngủ (bắt buộc khớp giới tính) ──
                # GV đã trực Ăn vẫn CÓ THỂ trực Ngủ (2 ca khác thời gian)
                # nhưng không được trực 2 phòng Ngủ cùng lúc
                for p in phongs_ngu:
                    req_gender = p.gioi_tinh  # ràng buộc nam/nữ
                    for _ in range(p.sl_diem_danh):
                        g = _pick(pool_dd, assigned_ngu, req_gender)
                        if g:
                            LichTrucCoDinh.objects.create(ma_phong=p, ma_gv=g, thu=thu)
                    for _ in range(p.sl_ho_tro):
                        g = _pick(pool_ht, assigned_ngu, req_gender)
                        if g:
                            LichTrucCoDinh.objects.create(ma_phong=p, ma_gv=g, thu=thu)
                            
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@require_POST
@role_required('admin', 'quan_ly')
def api_lichtruc_khung_save(request):
    try:
        data = json.loads(request.body)
        with transaction.atomic():
            LichTrucCoDinh.objects.all().delete()
            for row in data:
                p = Phong.objects.get(pk=row['ma_phong'])
                g = GiaoVien.objects.get(pk=row['ma_gv'])
                t = int(row['thu'])
                LichTrucCoDinh.objects.create(ma_phong=p, ma_gv=g, thu=t)
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@require_POST
@role_required('admin', 'quan_ly')
def api_lichtruc_khung_apply(request):
    # Apply to a specific week range (Not implemented yet - placeholder)
    return JsonResponse({'ok': True})


# ─────────────────────────────────────────────────────────
# API: Nạp Lịch Khung vào tuần cụ thể
# ─────────────────────────────────────────────────────────
@require_POST
@role_required('admin', 'quan_ly')
def api_apply_khung_to_week(request):
    """
    POST /api/lichtruc/apply-khung/
    Body: { tuan: 'YYYY-MM-DD', force: true|false }
    - tuan: ngày T2 của tuần
    - force: true = ghi đè ngày đã có | false = bỏ qua ngày đã có
    """
    try:
        body = json.loads(request.body)
        tuan_str = body.get('tuan')
        force    = body.get('force', False)
        mon = datetime.strptime(tuan_str, '%Y-%m-%d').date()

        created = 0
        skipped = 0
        overwritten = 0

        khung_all = list(
            LichTrucCoDinh.objects.all().select_related('ma_gv', 'ma_phong')
        )

        with transaction.atomic():
            for offset in range(5):  # T2=0 .. T6=4
                ngay = mon + timedelta(days=offset)
                thu  = offset  # 0=T2, 1=T3, ..., 4=T6
                khung_day = [k for k in khung_all if k.thu == thu]

                if not khung_day:
                    continue

                # Kiểm tra ngày này đã có phân công chưa
                existing_count = PhanCongTrucGV.objects.filter(ngay=ngay).count()

                if existing_count > 0 and not force:
                    skipped += existing_count
                    continue  # bỏ qua ngày đã có

                if existing_count > 0 and force:
                    deleted = PhanCongTrucGV.objects.filter(ngay=ngay).delete()
                    overwritten += existing_count

                for k in khung_day:
                    PhanCongTrucGV.objects.create(
                        ma_gv=k.ma_gv,
                        ma_phong=k.ma_phong,
                        ngay=ngay,
                        loai_truc=k.ma_phong.loai_phong,
                        xac_nhan_truc=True
                    )
                    created += 1

        return JsonResponse({'ok': True, 'created': created, 'skipped': skipped, 'overwritten': overwritten})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required(login_url='/login/')
def api_lichtruc_month(request):
    """
    GET /api/lichtruc/month/?thang=YYYY-MM
    Trả về tất cả PhanCongTrucGV trong tháng, kèm GV_LIST và PHONG_LIST
    """
    thang_str = request.GET.get('thang')
    try:
        year, month = map(int, thang_str.split('-'))
    except (ValueError, TypeError, AttributeError):
        today = date.today()
        year, month = today.year, today.month

    from calendar import monthrange
    last_day = monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year, month, last_day)

    qs = PhanCongTrucGV.objects.filter(
        ngay__gte=start, ngay__lte=end
    ).select_related('ma_gv', 'ma_phong', 'ma_gv_truc_thay')

    records = []
    for pc in qs:
        records.append({
            'id': pc.pk,
            'ma_gv_id': pc.ma_gv_id,
            'ma_gv_thay_id': pc.ma_gv_truc_thay_id,
            'ma_phong': pc.ma_phong.ma_phong,
            'ngay': pc.ngay.isoformat(),
            'loai_truc': pc.loai_truc,
            'xac_nhan': pc.xac_nhan_truc,
        })

    gv_list = list(GiaoVien.objects.filter(dang_lam=True).values('id', 'ho_ten', 'gioi_tinh'))
    phong_list = []
    for p in Phong.objects.all():
        phong_list.append({
            'ma': p.ma_phong,
            'loai': p.loai_phong,
            'gt': p.gioi_tinh,
        })

    return JsonResponse({
        'records': records,
        'gv_list': gv_list,
        'phong_list': phong_list,
        'thang': f'{year}-{month:02d}',
    })


@login_required(login_url='/login/')
def api_lichtruc_week_public(request):
    """
    GET /api/lichtruc/week-public/?tuan=YYYY-MM-DD
    Trả về tất cả PhanCongTrucGV trong tuần (T2–T7) cho tất cả role
    """
    tuan_str = request.GET.get('tuan')
    try:
        mon = datetime.strptime(tuan_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        mon = date.today()
        mon = mon - timedelta(days=mon.weekday())

    sun = mon + timedelta(days=6)
    qs = PhanCongTrucGV.objects.filter(
        ngay__gte=mon, ngay__lte=sun
    ).select_related('ma_gv', 'ma_phong', 'ma_gv_truc_thay')

    records = []
    for pc in qs:
        records.append({
            'id': pc.pk,
            'ma_gv_id': pc.ma_gv_id,
            'ma_gv_thay_id': pc.ma_gv_truc_thay_id,
            'ma_phong': pc.ma_phong.ma_phong,
            'ngay': pc.ngay.isoformat(),
            'loai_truc': pc.loai_truc,
            'xac_nhan': pc.xac_nhan_truc,
        })

    gv_list = list(GiaoVien.objects.filter(dang_lam=True).values('id', 'ho_ten', 'gioi_tinh', 'nhiem_vu'))
    phong_list = []
    for p in Phong.objects.all():
        phong_list.append({
            'ma': p.ma_phong,
            'loai': p.loai_phong,
            'gt': p.gioi_tinh,
        })

    ht = CauHinhHeThong.get()
    return JsonResponse({
        'records': records,
        'gv_list': gv_list,
        'phong_list': phong_list,
        'tuan': mon.isoformat(),
        'nam_hoc': ht.nam_hoc,
        'nguoi_phu_trach': ht.nguoi_phu_trach,
        'ten_truong': ht.ten_truong,
    })


@role_required('admin', 'quan_ly')
def api_lichtruc_week(request):
    """
    GET /api/lichtruc/week/?tuan=YYYY-MM-DD
    Trả về tất cả PhanCongTrucGV trong tuần (T2–T6)
    """
    tuan_str = request.GET.get('tuan')
    try:
        mon = datetime.strptime(tuan_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        mon = date.today()
        # Đi về ngày T2 của tuần hiện tại
        mon = mon - timedelta(days=mon.weekday())

    fri = mon + timedelta(days=4)
    qs = PhanCongTrucGV.objects.filter(
        ngay__gte=mon, ngay__lte=fri
    ).select_related('ma_gv', 'ma_phong', 'ma_gv_truc_thay')

    data = []
    for pc in qs:
        data.append({
            'id': pc.pk,
            'ma_gv_id': pc.ma_gv_id,
            'ma_gv_truc_thay_id': pc.ma_gv_truc_thay_id,
            'ma_phong_id': pc.ma_phong.ma_phong,
            'ngay': pc.ngay.isoformat(),
            'loai_truc': pc.loai_truc,
            'xac_nhan_truc': pc.xac_nhan_truc,
        })
    return JsonResponse({'records': data, 'tuan': mon.isoformat()})


# ─────────────────────────────────────────────────────────
# EXPORT EXCEL: Lịch ăn/ngủ GV 2 tuần
# ─────────────────────────────────────────────────────────
@role_required('admin', 'quan_ly')
def api_export_lich_gv(request):
    """
    GET /api/lichtruc/export/?tuan=YYYY-MM-DD
    Xuất Excel lịch ăn+ngủ của GV trong 2 tuần liên tiếp
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    tuan_str = request.GET.get('tuan')
    try:
        mon = datetime.strptime(tuan_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        mon = date.today() - timedelta(days=date.today().weekday())

    # 2 tuần = 10 ngày
    fri2 = mon + timedelta(days=11)  # T6 tuần 2
    all_dates = [mon + timedelta(days=i) for i in range(12) if (mon + timedelta(days=i)).weekday() < 5]

    phongs_an  = list(Phong.objects.filter(loai_phong=LoaiTruc.AN).order_by('ma_phong'))
    phongs_ngu = list(Phong.objects.filter(loai_phong=LoaiTruc.NGU).order_by('ma_phong'))
    all_phongs = phongs_an + phongs_ngu

    qs = PhanCongTrucGV.objects.filter(
        ngay__gte=mon, ngay__lte=fri2
    ).select_related('ma_gv', 'ma_phong', 'ma_gv_truc_thay')

    # Index: (ngay, ma_phong) -> list of ho_ten
    pc_map = {}
    for pc in qs:
        key = (pc.ngay.isoformat(), pc.ma_phong.ma_phong)
        if key not in pc_map:
            pc_map[key] = []
        name = pc.ma_gv_truc_thay.ho_ten if pc.ma_gv_truc_thay else pc.ma_gv.ho_ten
        pc_map[key].append(name)

    # ────────── Tạo workbook ──────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lịch trực GV'

    # Màu sắc
    CLR_HEADER  = 'FF1e3a5f'
    CLR_AN      = 'FFFef3c7'
    CLR_NGU     = 'FFede9fe'
    CLR_DAY_ODD = 'FFF8FAFF'
    CLR_DAY_EVEN= 'FFFFFFFF'
    CLR_WEEK2   = 'FFf0f9ff'

    thin = Side(style='thin', color='FFcbd5e1')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell_style(ws, row, col, value='', bold=False, bg=None, align='center', font_size=10, color='FF1e293b', wrap=False, italic=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=font_size, color=color, italic=italic, name='Arial')
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if bg:
            c.fill = PatternFill('solid', fgColor=bg)
        c.border = border_all
        return c

    row = 1
    # Tiêu đề
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + len(all_phongs))
    c = ws.cell(row=row, column=1, value='LỈCH TRỰC GIÁO VIÊN ĂN - NGỦ')
    c.font = Font(bold=True, size=14, color='FF1e3a5f', name='Arial')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 32
    row += 1

    # Kỳ xuất
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + len(all_phongs))
    c = ws.cell(row=row, column=1, value=f'Từ {mon.strftime("%d/%m/%Y")} đến {fri2.strftime("%d/%m/%Y")}')
    c.font = Font(size=10, italic=True, color='FF64748b', name='Arial')
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 18
    row += 1

    # Header hàng
    DAYS_VI = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6']
    cell_style(ws, row, 1, 'Ngày', bold=True, bg=CLR_HEADER, color='FFFFFFFF', font_size=11)
    ws.column_dimensions['A'].width = 18

    for i, p in enumerate(all_phongs):
        col = i + 2
        bg = CLR_AN if p.loai_phong == LoaiTruc.AN else CLR_NGU
        gt = '' if p.gioi_tinh is None else (' (Nam)' if p.gioi_tinh == 0 else ' (Nữ)')
        loai_txt = 'Ăn' if p.loai_phong == LoaiTruc.AN else 'Ngủ'
        cell_style(ws, row, col, f'P.{p.ma_phong}\n({loai_txt}{gt})', bold=True, bg=CLR_HEADER, color='FFFFFFFF', font_size=9, wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.row_dimensions[row].height = 40
    row += 1

    # Dữ liệu ngày
    for d_idx, ngay in enumerate(all_dates):
        wd = ngay.weekday()
        day_label = f'{DAYS_VI[wd]}\n{ngay.strftime("%d/%m/%Y")}'
        is_w2 = d_idx >= 5
        row_bg = CLR_WEEK2 if is_w2 else (CLR_DAY_ODD if d_idx % 2 == 0 else CLR_DAY_EVEN)

        cell_style(ws, row, 1, day_label, bold=True, bg=row_bg, align='left', wrap=True, font_size=9)
        for i, p in enumerate(all_phongs):
            col = i + 2
            gvs = pc_map.get((ngay.isoformat(), p.ma_phong), [])
            cell_style(ws, row, col, '\n'.join(gvs) if gvs else '—', bg=row_bg, wrap=True, font_size=9,
                       color='FF64748b' if not gvs else 'FF1e293b')
        ws.row_dimensions[row].height = 35
        row += 1

    # Freeze header
    ws.freeze_panes = ws.cell(row=4, column=2)

    # Output
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'lich-truc-GV_{mon.strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
